#!/usr/bin/env python3
"""listing-watch: config, roster load, name matching, prior-run diff, Excel + dashboard.

This module makes NO network calls. Page access happens through the browser, driven by the
skill, under a real person's own session. Everything here is local file work and arithmetic.

Subcommands:
    init    write config.json (and a roster template if the roster is missing)
    roster  validate the roster and emit the visit list as JSON
    report  match names, diff against the last run, write the workbook and dashboard

Requires Python 3.9+ and openpyxl.
"""

from __future__ import annotations

import argparse
import datetime as dt
import difflib
import hashlib
import html
import json
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

SITES = ("zillow", "redfin", "realtor", "homes")
SITE_LABEL = {
    "zillow": "Zillow",
    "redfin": "Redfin",
    "realtor": "realtor.com",
    "homes": "homes.com",
}

DEFAULT_QUALIFYING = ["PENDING", "CONTINGENT", "UNDER CONTRACT", "SOLD"]

# Page wording -> normalized status. Longest key wins, so order does not matter.
STATUS_MAP = {
    "pending": "PENDING",
    "pending sale": "PENDING",
    "sale pending": "PENDING",
    "pending continue to show": "PENDING",
    "contingent": "CONTINGENT",
    "active contingent": "CONTINGENT",
    "active under contract": "CONTINGENT",
    "under contract": "UNDER CONTRACT",
    "under agreement": "UNDER CONTRACT",
    "sold": "SOLD",
    "closed": "SOLD",
    "recently sold": "SOLD",
    "active": "ACTIVE",
    "for sale": "ACTIVE",
    "coming soon": "COMING SOON",
    "new construction": "ACTIVE",
    "off market": "OFF MARKET",
    "withdrawn": "OFF MARKET",
    "expired": "OFF MARKET",
    "canceled": "OFF MARKET",
    "cancelled": "OFF MARKET",
}

# Suffixes and decorations that appear on profile pages but are not part of a name.
NAME_NOISE = {
    "realtor", "realtors", "realtor®", "broker", "associate", "agent", "team",
    "pa", "p.a.", "llc", "inc", "gri", "abr", "crs", "srs", "epro", "e-pro",
    "mba", "jr", "sr", "ii", "iii", "iv", "the", "group", "sales", "salesperson",
    "licensed", "license", "realty",
}

NICKNAMES = {
    "bob": "robert", "rob": "robert", "bobby": "robert",
    "bill": "william", "will": "william", "billy": "william",
    "dick": "richard", "rick": "richard", "rich": "richard",
    "jim": "james", "jimmy": "james",
    "joe": "joseph", "joey": "joseph",
    "mike": "michael", "mikey": "michael",
    "tom": "thomas", "tommy": "thomas",
    "dave": "david", "davey": "david",
    "steve": "steven", "stephen": "steven",
    "chris": "christopher", "kit": "christopher",
    "chrissy": "christine", "christy": "christine", "tina": "christine",
    "kathy": "katherine", "kate": "katherine", "katie": "katherine",
    "cathy": "catherine",
    "liz": "elizabeth", "beth": "elizabeth", "betty": "elizabeth", "eliza": "elizabeth",
    "sue": "susan", "susie": "susan",
    "peggy": "margaret", "meg": "margaret", "maggie": "margaret",
    "nick": "nicholas", "tony": "anthony", "ed": "edward", "ted": "edward",
    "dan": "daniel", "danny": "daniel", "matt": "matthew", "greg": "gregory",
    "jeff": "jeffrey", "ken": "kenneth", "larry": "lawrence", "andy": "andrew",
    "drew": "andrew", "sam": "samuel", "ben": "benjamin", "alex": "alexander",
    "pat": "patricia", "patty": "patricia", "trish": "patricia",
    "debbie": "deborah", "deb": "deborah", "jen": "jennifer", "jenny": "jennifer",
    "sara": "sarah", "becky": "rebecca", "sandy": "sandra", "cindy": "cynthia",
}


# --------------------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------------------

def die(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(2)


def read_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def write_json(path: str, payload: Any) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


def require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        die("openpyxl is not installed. Run: pip install openpyxl")
    import openpyxl
    return openpyxl


def parse_date(value: Any) -> Optional[dt.date]:
    """Best-effort date parse. Returns None rather than raising, on purpose:
    a page that words a date oddly should not kill a run."""
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y", "%d %b %Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------------------
# name matching
# --------------------------------------------------------------------------------------

def normalize_name(raw: Optional[str]) -> str:
    if not raw:
        return ""
    text = raw.lower()
    text = text.split("|")[0].split("(")[0]
    # drop anything after a comma that looks like a credential list
    text = re.sub(r"[^a-z\s'.-]", " ", text)
    text = text.replace(".", " ").replace("-", " ").replace("'", "")
    tokens = [t for t in text.split() if t and t not in NAME_NOISE]
    tokens = [NICKNAMES.get(t, t) for t in tokens]
    return " ".join(sorted(tokens))


def match_names(roster_name: str, page_name: Optional[str], threshold: float) -> Tuple[str, float]:
    """Return (bucket, score). Buckets: exact, near, mismatch, unknown.

    `unknown` means the page gave us no name at all — that is not the agent's fault and not a
    mismatch, but it is not confirmation either, so it is kept as its own state.
    """
    if page_name is None or str(page_name).strip() == "":
        return "unknown", 0.0
    a, b = normalize_name(roster_name), normalize_name(page_name)
    if not a or not b:
        return "unknown", 0.0
    if a == b:
        return "exact", 1.0
    score = difflib.SequenceMatcher(None, a, b).ratio()
    # Surname agreement is a strong signal; a first-name variant should not read as a mismatch.
    tokens_a, tokens_b = set(a.split()), set(b.split())
    overlap = tokens_a & tokens_b
    if overlap and (len(overlap) >= max(len(tokens_a), len(tokens_b)) - 1):
        score = max(score, 0.90)
    score = round(score, 3)
    if score >= 0.97:
        return "exact", score
    if score >= threshold:
        return "near", score
    return "mismatch", score


def normalize_status(raw: Optional[str]) -> str:
    if not raw:
        return "OTHER"
    text = re.sub(r"[^a-z\s]", " ", str(raw).lower())
    text = " ".join(text.split())
    if text in STATUS_MAP:
        return STATUS_MAP[text]
    best = None
    for key, val in STATUS_MAP.items():
        if key in text and (best is None or len(key) > len(best[0])):
            best = (key, val)
    return best[1] if best else "OTHER"


# --------------------------------------------------------------------------------------
# init
# --------------------------------------------------------------------------------------

def cmd_init(args: argparse.Namespace) -> None:
    openpyxl = require_openpyxl()
    sites = [s.strip().lower() for s in args.sites.split(",") if s.strip()]
    bad = [s for s in sites if s not in SITES]
    if bad:
        die(f"unknown site(s): {', '.join(bad)}. Valid: {', '.join(SITES)}")

    config = {
        "schema_version": 1,
        "label": args.label,
        "roster_path": os.path.abspath(args.roster),
        "output_folder": os.path.abspath(args.output_folder),
        "sites": sites,
        "sold_window_days": args.sold_window_days,
        "qualifying_statuses": [s.strip().upper() for s in args.qualifying.split(",")]
        if args.qualifying else list(DEFAULT_QUALIFYING),
        "near_match_threshold": args.near_match_threshold,
        "owner": args.owner or "",
        "run_note": args.run_note or "",
        "constraints_note": args.constraints_note
        or "No outbound contact. Public profile pages only, no logged-in or MLS-gated data.",
        "created_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "last_run_at": None,
    }
    write_json(args.config, config)
    print(f"config written: {args.config}")

    work = os.path.join(os.path.dirname(os.path.abspath(args.config)))
    os.makedirs(os.path.join(work, "snapshots"), exist_ok=True)

    if not os.path.exists(config["roster_path"]):
        template = os.path.join(work, "roster_template.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"
        headers = ["agent_name", "agent_id", "zillow_url", "redfin_url",
                   "realtor_url", "homes_url", "notes"]
        ws.append(headers)
        ws.append(["Jane Example", "EX-001", "https://www.zillow.com/profile/example",
                   "", "https://www.realtor.com/realestateagents/example", "",
                   "delete this row before your first real run"])
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(16, len(h) + 4)
        ws.freeze_panes = "A2"
        wb.save(template)
        print(f"roster not found, template written: {template}")
        print("Fill it in, save it at the roster path above, then run the watch.")


# --------------------------------------------------------------------------------------
# roster
# --------------------------------------------------------------------------------------

def load_roster(config: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    openpyxl = require_openpyxl()
    path = config["roster_path"]
    if not os.path.exists(path):
        die(f"roster not found at {path}. Fix roster_path in the config, or create the file.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Roster"] if "Roster" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        die(f"roster at {path} is empty")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "agent_name" not in header:
        die("roster is missing an 'agent_name' column. See reference/output-contract.md")
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, key):
        i = idx.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    agents: List[Dict[str, Any]] = []
    issues: List[Dict[str, str]] = []
    seen: Dict[str, int] = {}

    for n, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = cell(row, "agent_name")
        if not name:
            issues.append({"row": str(n), "issue": "no agent_name", "detail": ""})
            continue
        urls = {}
        for site in config["sites"]:
            url = cell(row, f"{site}_url")
            if not url:
                continue
            if not url.lower().startswith(("http://", "https://")):
                issues.append({"row": str(n), "issue": f"{site}_url is not a URL",
                               "detail": f"{name}: {url[:80]}"})
                continue
            urls[site] = url
        if not urls:
            issues.append({"row": str(n), "issue": "no profile URLs for any in-scope site",
                           "detail": name})
        key = normalize_name(name)
        if key in seen:
            issues.append({"row": str(n), "issue": "duplicate agent name",
                           "detail": f"{name} also on row {seen[key]}"})
        else:
            seen[key] = n
        agents.append({
            "row": n,
            "agent_name": name,
            "agent_id": cell(row, "agent_id"),
            "notes": cell(row, "notes"),
            "urls": urls,
        })

    wb.close()
    return agents, issues


def cmd_roster(args: argparse.Namespace) -> None:
    config = read_json(args.config)
    agents, issues = load_roster(config)
    visits = []
    for a in agents:
        for site in config["sites"]:
            visits.append({
                "agent_name": a["agent_name"],
                "agent_id": a["agent_id"],
                "site": site,
                "profile_url": a["urls"].get(site, ""),
                "outcome": "url_missing" if site not in a["urls"] else None,
            })
    payload = {
        "generated_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "roster_path": config["roster_path"],
        "agent_count": len(agents),
        "visit_count": len([v for v in visits if v["profile_url"]]),
        "sites": config["sites"],
        "visits": visits,
        "roster_issues": issues,
    }
    out = args.out or os.path.join(os.path.dirname(os.path.abspath(args.config)), "roster.json")
    write_json(out, payload)
    print(f"roster: {len(agents)} agents, {payload['visit_count']} pages to visit, "
          f"{len(issues)} issue(s)")
    print(f"written: {out}")


# --------------------------------------------------------------------------------------
# report
# --------------------------------------------------------------------------------------

def hit_key(row: Dict[str, Any]) -> str:
    raw = "|".join([
        normalize_name(row["agent_name"]),
        row["site"],
        re.sub(r"\s+", " ", (row["address"] or "").strip().lower()),
        row["status"],
    ])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def build_rows(config, observations, period_start, period_end):
    threshold = float(config.get("near_match_threshold", 0.80))
    qualifying = set(config.get("qualifying_statuses", DEFAULT_QUALIFYING))
    window_days = int(config.get("sold_window_days", 14))
    sold_floor = period_end - dt.timedelta(days=window_days)

    hits, names, coverage = [], [], []
    per_site = {s: {"attempted": 0, "ok": 0, "blocked": 0, "page_changed": 0,
                    "url_missing": 0} for s in config["sites"]}
    unmapped = set()

    for v in observations.get("visits", []):
        site = str(v.get("site", "")).lower()
        outcome = v.get("outcome") or "url_missing"
        if site not in per_site:
            per_site[site] = {"attempted": 0, "ok": 0, "blocked": 0,
                              "page_changed": 0, "url_missing": 0}
        per_site[site][outcome if outcome in per_site[site] else "page_changed"] += 1
        if outcome != "url_missing":
            per_site[site]["attempted"] += 1

        bucket, score = match_names(v.get("agent_name", ""), v.get("name_on_page"), threshold)

        coverage.append({
            "agent_name": v.get("agent_name", ""),
            "agent_id": v.get("agent_id", ""),
            "site": SITE_LABEL.get(site, site),
            "outcome": outcome,
            "name_on_page": v.get("name_on_page") or "",
            "match": bucket,
            "match_score": score,
            "listings_seen": len(v.get("listings") or []),
            "profile_url": v.get("profile_url", ""),
            "note": v.get("note", "") or "",
        })

        if outcome == "ok" and bucket != "exact":
            names.append(coverage[-1])

        if outcome != "ok":
            continue

        for listing in v.get("listings") or []:
            raw_status = listing.get("status") or ""
            status = normalize_status(raw_status)
            if status == "OTHER" and raw_status:
                unmapped.add(raw_status.strip())
            if status not in qualifying:
                continue
            sdate = parse_date(listing.get("status_date"))
            if status == "SOLD":
                if sdate is None:
                    pass  # no date shown: keep it, flag it in the note rather than guess
                elif sdate < sold_floor:
                    continue
            hits.append({
                "agent_name": v.get("agent_name", ""),
                "agent_id": v.get("agent_id", ""),
                "name_on_page": v.get("name_on_page") or "",
                "match": bucket,
                "match_score": score,
                "site": site,
                "site_label": SITE_LABEL.get(site, site),
                "status": status,
                "status_raw": raw_status,
                "address": listing.get("address") or "",
                "status_date": sdate.isoformat() if sdate else "",
                "profile_url": v.get("profile_url", ""),
                "listing_url": listing.get("listing_url") or "",
                "notes": " ".join(x for x in [v.get("note") or "",
                                              "" if sdate or status != "SOLD"
                                              else "no sale date shown on page"] if x).strip(),
            })

    return hits, names, coverage, per_site, sorted(unmapped)


def load_previous(snapshot_dir: str) -> Tuple[set, Optional[str]]:
    if not os.path.isdir(snapshot_dir):
        return set(), None
    files = sorted(f for f in os.listdir(snapshot_dir) if f.endswith(".json"))
    if not files:
        return set(), None
    latest = files[-1]
    try:
        data = read_json(os.path.join(snapshot_dir, latest))
    except (json.JSONDecodeError, OSError):
        return set(), None
    return set(data.get("keys", [])), latest


def cmd_report(args: argparse.Namespace) -> None:
    openpyxl = require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill

    config = read_json(args.config)
    observations = read_json(args.observations)
    work = os.path.dirname(os.path.abspath(args.config))
    snapshot_dir = os.path.join(work, "snapshots")
    os.makedirs(snapshot_dir, exist_ok=True)

    period_end = parse_date(args.period_end) or dt.date.today()
    period_start = parse_date(args.period_start) or (
        period_end - dt.timedelta(days=int(config.get("sold_window_days", 14))))

    _, roster_issues = load_roster(config)
    hits, names, coverage, per_site, unmapped = build_rows(
        config, observations, period_start, period_end)

    previous, prev_file = load_previous(snapshot_dir)
    first_run = not previous
    for h in hits:
        h["key"] = hit_key(h)
        # No prior snapshot means everything is new, and the report says so plainly rather
        # than letting a first run read as a sudden burst of activity.
        h["new_since_last_run"] = "YES" if (first_run or h["key"] not in previous) else "no"

    order = {"mismatch": 0, "unknown": 1, "near": 2, "exact": 3}
    hits.sort(key=lambda h: (h["new_since_last_run"] != "YES",
                             order.get(h["match"], 9), h["agent_name"].lower()))

    attempted = sum(s["attempted"] for s in per_site.values())
    ok = sum(s["ok"] for s in per_site.values())
    blocked = sum(s["blocked"] for s in per_site.values())
    changed = sum(s["page_changed"] for s in per_site.values())
    skipped = sum(s["url_missing"] for s in per_site.values())
    coverage_pct = round(100.0 * ok / attempted, 1) if attempted else 0.0
    new_hits = sum(1 for h in hits if h["new_since_last_run"] == "YES")
    clean = blocked == 0 and changed == 0

    label = re.sub(r"[^A-Za-z0-9_.-]", "-", config.get("label") or "Listing-Watch")
    base = f"{label}_ListingWatch_{period_start.isoformat()}_to_{period_end.isoformat()}"
    out_folder = config["output_folder"]
    os.makedirs(out_folder, exist_ok=True)
    xlsx_path = os.path.join(out_folder, base + ".xlsx")
    html_path = os.path.join(out_folder, base + "_dashboard.html")

    # ---------------- workbook ----------------
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    red = PatternFill("solid", fgColor="F8CBAD")
    green = PatternFill("solid", fgColor="D9EAD3")
    amber = PatternFill("solid", fgColor="FFE599")

    def sheet(title, headers, rows, widths=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for i, _ in enumerate(headers, start=1):
            ws.cell(row=1, column=i).font = bold
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = (
                widths[i - 1] if widths and i - 1 < len(widths) else 20)
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    ws = wb.active
    ws.title = "Run summary"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60
    verdict = ("Complete run" if clean and coverage_pct == 100 else
               "INCOMPLETE RUN — see Coverage detail")
    summary = [
        ("Report", f"Listing watch, {config.get('label','')}"),
        ("Period covered", f"{period_start.isoformat()} to {period_end.isoformat()}"),
        ("SOLD window", f"{config.get('sold_window_days', 14)} days"),
        ("Generated", dt.datetime.now().astimezone().isoformat(timespec="seconds")),
        ("", ""),
        ("COVERAGE", verdict),
        ("Pages attempted", attempted),
        ("Pages read", ok),
        ("Blocked by site", blocked),
        ("Page changed / unrecognized", changed),
        ("Skipped, no URL on roster", skipped),
        ("Coverage", f"{coverage_pct}%"),
        ("", ""),
        ("Qualifying listings", len(hits)),
        ("New since last run", f"{new_hits}" + (" (first run, all are new)" if first_run else "")),
        ("Names to review", len(names)),
        ("Roster issues", len(roster_issues)),
        ("Compared against", prev_file or "nothing — this is the first run"),
    ]
    for k, v in summary:
        ws.append([k, v])
    ws["A6"].font = bold
    ws["B6"].fill = green if (clean and coverage_pct == 100) else red
    ws["B6"].font = bold
    if unmapped:
        ws.append(["", ""])
        ws.append(["Unrecognized status wording", "; ".join(unmapped[:20])])
        ws.cell(row=ws.max_row, column=2).fill = amber
    ws.append(["", ""])
    ws.append(["Per site", "attempted / read / blocked / changed / skipped"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for site, s in per_site.items():
        ws.append([SITE_LABEL.get(site, site),
                   f"{s['attempted']} / {s['ok']} / {s['blocked']} / "
                   f"{s['page_changed']} / {s['url_missing']}"])

    hits_ws = sheet(
        "Hits",
        ["agent_name", "agent_id", "name_on_page", "match", "match_score", "site",
         "status", "status_raw", "address", "status_date", "new_since_last_run",
         "profile_url", "listing_url", "notes"],
        [[h["agent_name"], h["agent_id"], h["name_on_page"], h["match"], h["match_score"],
          h["site_label"], h["status"], h["status_raw"], h["address"], h["status_date"],
          h["new_since_last_run"], h["profile_url"], h["listing_url"], h["notes"]]
         for h in hits],
        widths=[24, 12, 24, 10, 12, 14, 16, 18, 42, 13, 19, 44, 44, 30])
    for i, h in enumerate(hits, start=2):
        if h["new_since_last_run"] == "YES":
            hits_ws.cell(row=i, column=11).fill = green
        if h["match"] in ("mismatch", "unknown"):
            hits_ws.cell(row=i, column=4).fill = red
        elif h["match"] == "near":
            hits_ws.cell(row=i, column=4).fill = amber
    if not hits:
        hits_ws.append(["No qualifying listings found." if clean and coverage_pct == 100
                        else "No qualifying listings found, but coverage was incomplete "
                             "— see Run summary before treating this as a quiet night."])

    sheet("Names to review",
          ["agent_name", "agent_id", "site", "name_on_page", "match", "match_score",
           "listings_seen", "profile_url", "note"],
          [[n["agent_name"], n["agent_id"], n["site"], n["name_on_page"], n["match"],
            n["match_score"], n["listings_seen"], n["profile_url"], n["note"]] for n in names],
          widths=[24, 12, 14, 26, 10, 12, 14, 46, 30])

    cov_ws = sheet("Coverage detail",
                   ["agent_name", "agent_id", "site", "outcome", "name_on_page", "match",
                    "listings_seen", "profile_url", "note"],
                   [[c["agent_name"], c["agent_id"], c["site"], c["outcome"],
                     c["name_on_page"], c["match"], c["listings_seen"], c["profile_url"],
                     c["note"]] for c in coverage],
                   widths=[24, 12, 14, 14, 26, 10, 14, 46, 30])
    for i, c in enumerate(coverage, start=2):
        if c["outcome"] in ("blocked", "page_changed"):
            cov_ws.cell(row=i, column=4).fill = red

    sheet("Roster issues", ["row", "issue", "detail"],
          [[r["row"], r["issue"], r["detail"]] for r in roster_issues],
          widths=[8, 40, 60])

    wb.save(xlsx_path)

    # ---------------- dashboard ----------------
    def esc(x):
        return html.escape(str(x if x is not None else ""))

    banner_class = "ok" if (clean and coverage_pct == 100) else "bad"
    banner_text = (f"Coverage {coverage_pct}% — all {attempted} pages read"
                   if clean and coverage_pct == 100 else
                   f"Coverage {coverage_pct}% — {blocked} blocked, {changed} changed, "
                   f"{skipped} skipped. Check those agents by hand today.")

    def table(headers, rows):
        head = "".join(f"<th>{esc(h)}</th>" for h in headers)
        if not rows:
            return (f"<table><thead><tr>{head}</tr></thead><tbody><tr>"
                    f"<td colspan='{len(headers)}' class='empty'>Nothing here.</td>"
                    f"</tr></tbody></table>")
        body = "".join("<tr>" + "".join(f"<td>{esc(c)}</td>" for c in r) + "</tr>"
                       for r in rows)
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    hit_rows = [[h["new_since_last_run"], h["agent_name"], h["name_on_page"], h["match"],
                 h["site_label"], h["status"], h["address"], h["status_date"]] for h in hits]
    name_rows = [[n["agent_name"], n["site"], n["name_on_page"], n["match"],
                  n["match_score"], n["outcome"]] for n in names]
    cov_rows = [[c["agent_name"], c["site"], c["outcome"], c["listings_seen"]]
                for c in coverage if c["outcome"] != "ok"]

    doc = f"""<!DOCTYPE html>
<meta charset="utf-8"><title>Listing watch {esc(period_start)} to {esc(period_end)}</title>
<style>
 body{{font:15px/1.5 -apple-system,Segoe UI,Roboto,sans-serif;margin:0;padding:28px;
      background:#f6f7f9;color:#1a1a1a}}
 h1{{font-size:21px;margin:0 0 4px}} h2{{font-size:16px;margin:30px 0 8px}}
 .sub{{color:#666;margin-bottom:18px}}
 .banner{{padding:14px 18px;border-radius:8px;font-weight:600;margin-bottom:20px}}
 .banner.ok{{background:#d9ead3;border:1px solid #9fc78a}}
 .banner.bad{{background:#f8cbad;border:1px solid #d98b5f}}
 .cards{{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:8px}}
 .card{{background:#fff;border:1px solid #e2e4e8;border-radius:8px;padding:14px 18px;min-width:130px}}
 .card .n{{font-size:26px;font-weight:650}} .card .l{{color:#666;font-size:13px}}
 table{{border-collapse:collapse;width:100%;background:#fff;border:1px solid #e2e4e8;
        border-radius:8px;overflow:hidden;font-size:13.5px}}
 th{{background:#eef0f3;text-align:left;padding:8px 10px;font-weight:600}}
 td{{padding:7px 10px;border-top:1px solid #eef0f3;vertical-align:top}}
 td.empty{{color:#888;text-align:center;padding:18px}}
 tr:hover td{{background:#fafbfc}}
 .foot{{margin-top:30px;color:#777;font-size:12.5px}}
</style>
<h1>Listing watch</h1>
<div class="sub">{esc(config.get('label',''))} &middot; {esc(period_start)} to {esc(period_end)}
 &middot; generated {esc(dt.datetime.now().astimezone().strftime('%Y-%m-%d %H:%M'))}</div>
<div class="banner {banner_class}">{esc(banner_text)}</div>
<div class="cards">
 <div class="card"><div class="n">{len(hits)}</div><div class="l">qualifying listings</div></div>
 <div class="card"><div class="n">{new_hits}</div><div class="l">new since last run</div></div>
 <div class="card"><div class="n">{len(names)}</div><div class="l">names to review</div></div>
 <div class="card"><div class="n">{ok}/{attempted}</div><div class="l">pages read</div></div>
</div>
{"<p class='sub'>First run, so every listing is flagged new.</p>" if first_run else ""}
<h2>Hits</h2>
{table(["New","Agent","Name on page","Match","Site","Status","Address","Date"], hit_rows)}
<h2>Names to review</h2>
{table(["Agent","Site","Name on page","Match","Score","Outcome"], name_rows)}
<h2>Pages not read</h2>
{table(["Agent","Site","Outcome","Listings seen"], cov_rows)}
<div class="foot">Public profile pages only. No outbound contact was made and none is
 suggested by this report. Compared against: {esc(prev_file or 'nothing, first run')}.</div>
"""
    with open(html_path, "w", encoding="utf-8") as fh:
        fh.write(doc)

    # ---------------- snapshot ----------------
    if attempted and coverage_pct > 0:
        stamp = dt.datetime.now().strftime("%Y-%m-%dT%H%M%S")
        write_json(os.path.join(snapshot_dir, f"{stamp}.json"), {
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "coverage_pct": coverage_pct,
            "keys": sorted({h["key"] for h in hits}),
        })

    config["last_run_at"] = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    write_json(args.config, config)

    print(f"COVERAGE {coverage_pct}% — {ok} read, {blocked} blocked, {changed} changed, "
          f"{skipped} skipped")
    print(f"HITS {len(hits)} ({new_hits} new) | NAMES TO REVIEW {len(names)} | "
          f"ROSTER ISSUES {len(roster_issues)}")
    if unmapped:
        print("UNRECOGNIZED STATUS WORDING: " + "; ".join(unmapped[:10]))
    print(xlsx_path)
    print(html_path)


# --------------------------------------------------------------------------------------

def main() -> None:
    p = argparse.ArgumentParser(prog="lw.py", description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)

    i = sub.add_parser("init", help="write config.json and a roster template if needed")
    i.add_argument("--config", required=True)
    i.add_argument("--roster", required=True)
    i.add_argument("--output-folder", required=True)
    i.add_argument("--sites", default=",".join(SITES))
    i.add_argument("--sold-window-days", type=int, default=14)
    i.add_argument("--near-match-threshold", type=float, default=0.80)
    i.add_argument("--qualifying", default="")
    i.add_argument("--label", default="Listing-Watch")
    i.add_argument("--owner", default="")
    i.add_argument("--run-note", default="")
    i.add_argument("--constraints-note", default="")
    i.set_defaults(func=cmd_init)

    r = sub.add_parser("roster", help="validate the roster and emit the visit list")
    r.add_argument("--config", required=True)
    r.add_argument("--out", default="")
    r.set_defaults(func=cmd_roster)

    rep = sub.add_parser("report", help="match, diff, and write the workbook and dashboard")
    rep.add_argument("--config", required=True)
    rep.add_argument("--observations", required=True)
    rep.add_argument("--period-start", default="")
    rep.add_argument("--period-end", default="")
    rep.set_defaults(func=cmd_report)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
